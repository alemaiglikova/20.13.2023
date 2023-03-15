import psycopg2
from openpyxl import load_workbook


conn = psycopg2.connect(
    host="hostname",
    database="database_name",
    user="username",
    password="password"
)

wb = load_workbook('data.xlsx')
sheet = wb.active

for row in sheet.iter_rows(min_row=2, values_only=True):
    id, title, description, success, deadline, data_created = row
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO mytable (id, title, description, success, deadline, data_created) VALUES (%s, %s, %s, %s, %s, %s)",
        (id, title, description, success, deadline, data_created)
    )
    conn.commit()

conn.close()